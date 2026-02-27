"""
ExcelExtractor: extracts tabular data from .xlsx / .xls files.
"""

from typing import List
from core.models import DocumentChunk, FileType
from .base_extractor import BaseExtractor


class ExcelExtractor(BaseExtractor):
    """Extracts sheet data from Excel files as readable text."""

    SUPPORTED_EXTENSIONS = ["xlsx", "xls"]

    def extract(self, file_path: str) -> List[DocumentChunk]:
        chunks: List[DocumentChunk] = []
        source = file_path.split("/")[-1]
        chunk_idx = 0

        try:
            import openpyxl

            wb = openpyxl.load_workbook(file_path, data_only=True)

            for sheet_num, sheet_name in enumerate(wb.sheetnames, start=1):
                ws = wb[sheet_name]
                rows_text = []

                for row in ws.iter_rows(values_only=True):
                    cells = [str(cell) if cell is not None else "" for cell in row]
                    if any(c.strip() for c in cells):
                        rows_text.append(" | ".join(cells))

                if rows_text:
                    sheet_text = f"[Sheet: {sheet_name}]\n" + "\n".join(rows_text)
                    chunks.append(self._make_text_chunk(
                        text=sheet_text,
                        source=source,
                        page=sheet_num,
                        chunk_index=chunk_idx,
                        file_type=FileType.EXCEL,
                    ))
                    chunk_idx += 1

        except ImportError as e:
            raise RuntimeError(
                f"Missing dependency for Excel extraction: {e}. "
                "Run: pip install openpyxl"
            )
        except Exception as e:
            raise RuntimeError(f"Excel extraction failed for '{file_path}': {e}")

        return chunks